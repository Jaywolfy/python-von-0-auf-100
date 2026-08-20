"""
╔══════════════════════════════════════════════════════════════════╗
║  MODUL 24 · MUSTERLÖSUNGEN — Excel                               ║
╚══════════════════════════════════════════════════════════════════╝
"""
from pathlib import Path
import random

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ pip install openpyxl")
    raise SystemExit(0)

LOES = Path(__file__).parent / "_loesung24"
LOES.mkdir(exist_ok=True)
random.seed(7)

KOPF_BLAU = "4472C4"
GRUEN, GELB, ROT = "C6EFCE", "FFEB9C", "FFC7CE"

# Testdatei erzeugen
_m = Workbook()
_b = _m.active
_b.title = "Mitarbeiter"
_b.append(["Name", "Abteilung", "Eintritt", "Gehalt", "Stunden"])
PERSONAL = [
    ("Anna Müller", "IT", 2019, 4800, 40),
    ("Bernd Schmidt", "Marketing", 2021, 3900, 32),
    ("Clara Weiß", "IT", 2017, 5600, 40),
    ("David Braun", "Vertrieb", 2022, 4200, 40),
    ("Eva Krüger", "IT", 2020, 5100, 30),
    ("Felix Wagner", "Marketing", 2018, 4400, 40),
    ("Greta Hoff", "Vertrieb", 2023, 3700, 25),
    ("Hans Berg", "IT", 2015, 6200, 40),
]
for z in PERSONAL:
    _b.append(list(z))
QUELLE = LOES / "personal.xlsx"
_m.save(QUELLE)


print("=" * 62, "\nAUFGABE 1 🟢\n", "=" * 62)

mappe = load_workbook(QUELLE)
blatt = mappe.active

print(f"  a) Blätter:  {mappe.sheetnames}")
print(f"  b) Größe:    {blatt.max_row} Zeilen x {blatt.max_column} Spalten")
print(f"  c) A1:       {blatt['A1'].value}")
namen = [blatt.cell(row=r, column=1).value for r in range(2, blatt.max_row + 1)]
print(f"  d) Namen:    {namen}")


print("\n" + "=" * 62, "\nAUFGABE 2 🟢\n", "=" * 62)

print(f"  {'Name':<16}{'Abteilung':<12}{'Eintritt':>9}{'Gehalt':>10}{'Std':>6}")
print("  " + "-" * 53)
for name, abt, eintritt, gehalt, stunden in blatt.iter_rows(min_row=2, values_only=True):
    print(f"  {name:<16}{abt:<12}{eintritt:>9}{gehalt:>10,}{stunden:>6}")


print("\n" + "=" * 62, "\nAUFGABE 3 🟡\n", "=" * 62)

mitarbeiter = [
    {"name": n, "abteilung": a, "eintritt": e, "gehalt": g, "stunden": s}
    for n, a, e, g, s in blatt.iter_rows(min_row=2, values_only=True)
]

gehaelter = [m["gehalt"] for m in mitarbeiter]
print(f"  a) Gesamtkosten:   {sum(gehaelter):>10,} € / Monat")
print(f"  b) Ø Gehalt:       {sum(gehaelter) / len(gehaelter):>10,.2f} €")

pro_abteilung = {}
for m in mitarbeiter:
    pro_abteilung.setdefault(m["abteilung"], []).append(m["gehalt"])

print("  c) Pro Abteilung:")
for abt, werte in sorted(pro_abteilung.items()):
    print(f"       {abt:<12}{len(werte)} MA  Summe {sum(werte):>7,} €  "
          f"Ø {sum(werte) / len(werte):>8,.0f} €")

top = max(mitarbeiter, key=lambda m: m["gehalt"])
flop = min(mitarbeiter, key=lambda m: m["gehalt"])
print(f"  d) Höchstes:       {top['name']} ({top['gehalt']:,} €)")
print(f"     Niedrigstes:    {flop['name']} ({flop['gehalt']:,} €)")

teilzeit = [m for m in mitarbeiter if m["stunden"] < 40]
print(f"  e) Teilzeit:       {len(teilzeit)} ({[m['name'] for m in teilzeit]})")

print("  f) Stundenlohn:")
for m in sorted(mitarbeiter, key=lambda m: -m["gehalt"] / (m["stunden"] * 4.33)):
    lohn = m["gehalt"] / (m["stunden"] * 4.33)
    print(f"       {m['name']:<16}{lohn:>7.2f} €/h")


print("\n" + "=" * 62, "\nAUFGABE 4 🟡\n", "=" * 62)


def formatiere_kopfzeile(arbeitsblatt, farbe=KOPF_BLAU):
    """Formatiert die erste Zeile als Kopfzeile."""
    for zelle in arbeitsblatt[1]:
        if zelle.value is not None:
            zelle.font = Font(bold=True, color="FFFFFF")
            zelle.fill = PatternFill("solid", fgColor=farbe)
            zelle.alignment = Alignment(horizontal="center")
    arbeitsblatt.row_dimensions[1].height = 22
    arbeitsblatt.freeze_panes = "A2"


def breite_anpassen(arbeitsblatt, zusatz=4):
    """Passt alle Spaltenbreiten an den Inhalt an."""
    for spalte in range(1, arbeitsblatt.max_column + 1):
        laengste = max(len(str(arbeitsblatt.cell(row=r, column=spalte).value or ""))
                       for r in range(1, arbeitsblatt.max_row + 1))
        arbeitsblatt.column_dimensions[get_column_letter(spalte)].width = laengste + zusatz


aus_mappe = Workbook()
aus_blatt = aus_mappe.active
aus_blatt.title = "Abteilungen"
aus_blatt.append(["Abteilung", "Anzahl", "Gesamtgehalt", "Ø Gehalt"])

for abt, werte in sorted(pro_abteilung.items(), key=lambda p: -sum(p[1])):
    aus_blatt.append([abt, len(werte), sum(werte),
                      round(sum(werte) / len(werte), 2)])

for r in range(2, aus_blatt.max_row + 1):
    aus_blatt.cell(row=r, column=3).number_format = '#,##0.00 "€"'
    aus_blatt.cell(row=r, column=4).number_format = '#,##0.00 "€"'

formatiere_kopfzeile(aus_blatt)
breite_anpassen(aus_blatt)

AUSWERTUNG = LOES / "auswertung.xlsx"
aus_mappe.save(AUSWERTUNG)
print(f"  ✅ {AUSWERTUNG.name} erzeugt ({aus_blatt.max_row - 1} Abteilungen)")


print("\n" + "=" * 62, "\nAUFGABE 5 🟡\n", "=" * 62)

mappe = load_workbook(QUELLE)
blatt = mappe.active

blatt["F1"] = "Jahre dabei"
blatt["G1"] = "Jahresgehalt"

for r in range(2, blatt.max_row + 1):
    eintritt = blatt.cell(row=r, column=3).value
    blatt.cell(row=r, column=6, value=2026 - eintritt)
    z = blatt.cell(row=r, column=7, value=f"=D{r}*12")     # ⭐ Excel-Formel
    z.number_format = '#,##0.00 "€"'

formatiere_kopfzeile(blatt)
breite_anpassen(blatt)

ERWEITERT = LOES / "personal_erweitert.xlsx"
mappe.save(ERWEITERT)          # ⚠️ NEUER Name - Original bleibt heil
print(f"  ✅ {ERWEITERT.name} (Original unverändert ✅)")


print("\n" + "=" * 62, "\nAUFGABE 6 🔴\n", "=" * 62)

mappe = load_workbook(QUELLE)
blatt = mappe.active

for r in range(2, blatt.max_row + 1):
    gehalt = blatt.cell(row=r, column=4).value
    if gehalt > 5000:
        farbe = GRUEN
    elif gehalt < 4000:
        farbe = ROT
    else:
        farbe = GELB
    for c in range(1, blatt.max_column + 1):
        blatt.cell(row=r, column=c).fill = PatternFill("solid", fgColor=farbe)
    blatt.cell(row=r, column=4).number_format = '#,##0 "€"'

formatiere_kopfzeile(blatt)
breite_anpassen(blatt)
blatt.auto_filter.ref = blatt.dimensions       # ⭐ Autofilter

UEBERSICHT = LOES / "gehaltsuebersicht.xlsx"
mappe.save(UEBERSICHT)
print(f"  ✅ {UEBERSICHT.name}")
print("     🟢 > 5000 €   🟡 4000-5000 €   🔴 < 4000 €   + Autofilter")


print("\n" + "=" * 62, "\nAUFGABE 7 🔴\n", "=" * 62)

PRODUKTE = ["Laptop", "Maus", "Monitor", "Tastatur", "Headset"]

for q in range(1, 5):
    m = Workbook()
    b = m.active
    b.title = "Umsatz"
    b.append(["Produkt", "Umsatz"])
    for p in PRODUKTE:
        b.append([p, round(random.uniform(1000, 20000), 2)])
    m.save(LOES / f"quartal_{q}.xlsx")
print(f"  a) 4 Quartalsdateien erzeugt")

sammlung = Workbook()
ziel = sammlung.active
ziel.title = "Alle Quartale"
ziel.append(["Quartal", "Produkt", "Umsatz"])

pro_quartal = {}
for pfad in sorted(LOES.glob("quartal_*.xlsx")):
    quartal = f"Q{pfad.stem.split('_')[1]}"
    quelle = load_workbook(pfad, data_only=True)["Umsatz"]
    for produkt, umsatz in quelle.iter_rows(min_row=2, values_only=True):
        ziel.append([quartal, produkt, umsatz])
        pro_quartal[quartal] = pro_quartal.get(quartal, 0) + umsatz

for r in range(2, ziel.max_row + 1):
    ziel.cell(row=r, column=3).number_format = '#,##0.00 "€"'
formatiere_kopfzeile(ziel, "70AD47")
breite_anpassen(ziel)

zusammen = sammlung.create_sheet("Zusammenfassung")
zusammen.append(["Quartal", "Umsatz"])
for quartal, umsatz in sorted(pro_quartal.items()):
    zusammen.append([quartal, round(umsatz, 2)])
for r in range(2, zusammen.max_row + 1):
    zusammen.cell(row=r, column=2).number_format = '#,##0.00 "€"'
formatiere_kopfzeile(zusammen, "70AD47")
breite_anpassen(zusammen)

JAHR = LOES / "jahresuebersicht.xlsx"
sammlung.save(JAHR)
print(f"  b) ✅ {JAHR.name}: {ziel.max_row - 1} Zeilen aus 4 Dateien")
print("  c) ✅ Blatt 'Zusammenfassung':")
for quartal, umsatz in sorted(pro_quartal.items()):
    print(f"       {quartal}  {umsatz:>12,.2f} €  {'█' * int(umsatz / 5000)}")


print("\n" + "=" * 62, "\nAUFGABE 8 ⭐\n", "=" * 62)

mappe = load_workbook(AUSWERTUNG)
blatt = mappe["Abteilungen"]

diagramm = BarChart()
diagramm.type = "col"
diagramm.title = "Gehaltskosten pro Abteilung"
diagramm.y_axis.title = "Euro"
diagramm.x_axis.title = "Abteilung"
diagramm.height, diagramm.width = 8, 14

werte = Reference(blatt, min_col=3, min_row=1, max_row=blatt.max_row)
kategorien = Reference(blatt, min_col=1, min_row=2, max_row=blatt.max_row)
diagramm.add_data(werte, titles_from_data=True)
diagramm.set_categories(kategorien)
blatt.add_chart(diagramm, "F2")

mappe.save(AUSWERTUNG)
print(f"  ✅ Balkendiagramm in {AUSWERTUNG.name} eingefügt")

print(f"""
{'=' * 62}
  📁 Alle Dateien liegen in: {LOES}
  Öffne sie in Excel/LibreOffice und schau sie dir an! 👀

🎉 Modul 24 geschafft! Das ist der Zeitfresser Nr. 1 im Büro - erledigt. 📊
""")
