"""
Modul 24 - Beispiel: Excel automatisieren mit openpyxl

Erzeugt sich alle Testdateien selbst. Nichts geht kaputt. 😌
"""
from pathlib import Path
import random

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl fehlt. Installieren mit:\n   pip install openpyxl\n")
    raise SystemExit(0)

ORDNER = Path(__file__).parent / "_excel24"
ORDNER.mkdir(exist_ok=True)

random.seed(42)

# ====================================================================
print("=" * 62, "\n1. EINE EXCEL-DATEI ERZEUGEN\n", "=" * 62)

mappe = Workbook()
blatt = mappe.active
blatt.title = "Rohdaten"

kopf = ["Monat", "Produkt", "Kategorie", "Menge", "Einzelpreis", "Umsatz"]
blatt.append(kopf)

MONATE = ["Januar", "Februar", "März", "April", "Mai", "Juni"]
PRODUKTE = [("Laptop", "Technik", 899.99), ("Maus", "Technik", 25.50),
            ("Schreibtisch", "Möbel", 349.00), ("Bürostuhl", "Möbel", 259.90),
            ("Monitor", "Technik", 219.00)]

zeile_nr = 2
for monat in MONATE:
    for produkt, kategorie, preis in PRODUKTE:
        menge = random.randint(1, 15)
        blatt.append([monat, produkt, kategorie, menge, preis,
                      f"=D{zeile_nr}*E{zeile_nr}"])       # ⭐ Excel-Formel!
        zeile_nr += 1

datei = ORDNER / "verkaeufe.xlsx"
mappe.save(datei)
print(f"  ✅ {datei.name} erzeugt: {blatt.max_row - 1} Datenzeilen")

# ====================================================================
print("\n" + "=" * 62, "\n2. FORMATIEREN 🎨\n", "=" * 62)

mappe = load_workbook(datei)
blatt = mappe["Rohdaten"]

KOPF_FARBE = "4472C4"
duenn = Side(style="thin", color="B0B0B0")
rahmen = Border(left=duenn, right=duenn, top=duenn, bottom=duenn)

# Kopfzeile
for zelle in blatt[1]:
    zelle.font = Font(bold=True, size=11, color="FFFFFF")
    zelle.fill = PatternFill("solid", fgColor=KOPF_FARBE)
    zelle.alignment = Alignment(horizontal="center", vertical="center")
    zelle.border = rahmen

blatt.row_dimensions[1].height = 24
blatt.freeze_panes = "A2"                 # Kopfzeile beim Scrollen fixieren

# Spaltenbreiten automatisch
for spalte in range(1, blatt.max_column + 1):
    buchstabe = get_column_letter(spalte)
    laengste = max(len(str(blatt.cell(row=r, column=spalte).value or ""))
                   for r in range(1, blatt.max_row + 1))
    blatt.column_dimensions[buchstabe].width = laengste + 4

# Zahlenformate + Zebrastreifen
for r in range(2, blatt.max_row + 1):
    blatt.cell(row=r, column=5).number_format = '#,##0.00 "€"'
    blatt.cell(row=r, column=6).number_format = '#,##0.00 "€"'
    if r % 2 == 0:
        for c in range(1, blatt.max_column + 1):
            blatt.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")

# Summenzeile
summen_zeile = blatt.max_row + 2
blatt.cell(row=summen_zeile, column=1, value="GESAMT").font = Font(bold=True, size=12)
summe = blatt.cell(row=summen_zeile, column=6,
                   value=f"=SUM(F2:F{blatt.max_row})")
summe.font = Font(bold=True, size=12)
summe.number_format = '#,##0.00 "€"'
summe.fill = PatternFill("solid", fgColor="FFF2CC")

mappe.save(datei)
print("  ✅ Kopfzeile blau + fett, Zebrastreifen, Eurоformat,")
print("     Spaltenbreiten, fixierte Kopfzeile, Summenformel")

# ====================================================================
print("\n" + "=" * 62, "\n3. LESEN & AUSWERTEN\n", "=" * 62)

# ⚠️ data_only=True liefert Werte statt Formeln.
# Hinweis: Die Formelwerte existieren erst, NACHDEM Excel die Datei
# einmal geöffnet und gespeichert hat. Deshalb rechnen wir hier selbst.
mappe = load_workbook(datei)
blatt = mappe["Rohdaten"]

daten = []
for monat, produkt, kategorie, menge, preis, _ in blatt.iter_rows(
        min_row=2, max_row=31, values_only=True):
    if monat is None:
        continue
    daten.append({"monat": monat, "produkt": produkt, "kategorie": kategorie,
                  "menge": menge, "preis": preis,
                  "umsatz": round(menge * preis, 2)})

gesamt = sum(d["umsatz"] for d in daten)
print(f"  {len(daten)} Datensätze gelesen")
print(f"  Gesamtumsatz: {gesamt:,.2f} €".replace(",", "."))

pro_monat = {}
pro_kategorie = {}
for d in daten:
    pro_monat[d["monat"]] = pro_monat.get(d["monat"], 0) + d["umsatz"]
    pro_kategorie[d["kategorie"]] = pro_kategorie.get(d["kategorie"], 0) + d["umsatz"]

hoechster = max(pro_monat.values())
print("\n  Umsatz pro Monat:")
for monat in MONATE:
    wert = pro_monat[monat]
    balken = "█" * int(wert / hoechster * 30)
    print(f"    {monat:<10}{wert:>12,.2f} €  {balken}")

print("\n  Umsatz pro Kategorie:")
for kategorie, wert in sorted(pro_kategorie.items(), key=lambda p: -p[1]):
    print(f"    {kategorie:<10}{wert:>12,.2f} €  ({wert / gesamt:.1%})")

# ====================================================================
print("\n" + "=" * 62, "\n4. AUSWERTUNGSBLATT + DIAGRAMM 📈\n", "=" * 62)

auswertung = mappe.create_sheet("Auswertung", 0)      # als erstes Blatt

auswertung["A1"] = "MONATSAUSWERTUNG 2026"
auswertung["A1"].font = Font(bold=True, size=16, color=KOPF_FARBE)
auswertung.merge_cells("A1:C1")

auswertung["A3"] = "Monat"
auswertung["B3"] = "Umsatz"
auswertung["C3"] = "Anteil"
for zelle in auswertung[3]:
    if zelle.value:
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.fill = PatternFill("solid", fgColor=KOPF_FARBE)

for i, monat in enumerate(MONATE, start=4):
    auswertung.cell(row=i, column=1, value=monat)
    z = auswertung.cell(row=i, column=2, value=round(pro_monat[monat], 2))
    z.number_format = '#,##0.00 "€"'
    a = auswertung.cell(row=i, column=3, value=pro_monat[monat] / gesamt)
    a.number_format = "0.0%"

auswertung.column_dimensions["A"].width = 16
auswertung.column_dimensions["B"].width = 16
auswertung.column_dimensions["C"].width = 12

diagramm = BarChart()
diagramm.title = "Umsatz pro Monat"
diagramm.y_axis.title = "Euro"
diagramm.x_axis.title = "Monat"
diagramm.height = 8
diagramm.width = 16
werte = Reference(auswertung, min_col=2, min_row=3, max_row=3 + len(MONATE))
kategorien = Reference(auswertung, min_col=1, min_row=4, max_row=3 + len(MONATE))
diagramm.add_data(werte, titles_from_data=True)
diagramm.set_categories(kategorien)
auswertung.add_chart(diagramm, "E3")

mappe.save(datei)
print(f"  ✅ Blatt 'Auswertung' mit Balkendiagramm hinzugefügt")
print(f"  📄 Blätter in der Datei: {mappe.sheetnames}")

# ====================================================================
print("\n" + "=" * 62, "\n5. MEHRERE DATEIEN ZUSAMMENFÜHREN 🔗\n", "=" * 62)

# Erst drei "Filial-Dateien" erzeugen
for filiale in ("Berlin", "Hamburg", "München"):
    m = Workbook()
    b = m.active
    b.title = "Umsatz"
    b.append(["Produkt", "Umsatz"])
    for produkt, _, _ in PRODUKTE:
        b.append([produkt, round(random.uniform(500, 5000), 2)])
    m.save(ORDNER / f"filiale_{filiale}.xlsx")

print(f"  3 Filialdateien erzeugt")

# Dann zusammenführen
sammlung = Workbook()
ziel = sammlung.active
ziel.title = "Alle Filialen"
ziel.append(["Filiale", "Produkt", "Umsatz"])

for pfad in sorted(ORDNER.glob("filiale_*.xlsx")):
    filiale = pfad.stem.replace("filiale_", "")
    quelle = load_workbook(pfad, data_only=True)["Umsatz"]
    for produkt, umsatz in quelle.iter_rows(min_row=2, values_only=True):
        ziel.append([filiale, produkt, umsatz])

for zelle in ziel[1]:
    zelle.font = Font(bold=True, color="FFFFFF")
    zelle.fill = PatternFill("solid", fgColor="70AD47")
for r in range(2, ziel.max_row + 1):
    ziel.cell(row=r, column=3).number_format = '#,##0.00 "€"'
for spalte, breite in (("A", 14), ("B", 18), ("C", 14)):
    ziel.column_dimensions[spalte].width = breite

zusammen = ORDNER / "alle_filialen.xlsx"
sammlung.save(zusammen)
print(f"  ✅ {zusammen.name}: {ziel.max_row - 1} Zeilen aus 3 Dateien")

print(f"""
  💡 GENAU DAS ist der Alltagsnutzen:
     40 Filialdateien einsammeln, zusammenführen, formatieren,
     Diagramm dazu - in unter einer Sekunde. Jeden Monat. 🎉

  📁 Alle Dateien liegen in: {ORDNER}
""")
