"""
╔══════════════════════════════════════════════════════════════════╗
║  MODUL 24 · AUFGABEN — Excel automatisieren                      ║
║  pip install openpyxl                                            ║
║  Alle Testdateien werden automatisch erzeugt. 😌                  ║
╚══════════════════════════════════════════════════════════════════╝
"""
from pathlib import Path
import random

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("❌ pip install openpyxl")
    raise SystemExit(0)

UEB = Path(__file__).parent / "_uebung24"
UEB.mkdir(exist_ok=True)
random.seed(7)

# --- Testdatei erzeugen (nicht ändern) ------------------------------
_mappe = Workbook()
_b = _mappe.active
_b.title = "Mitarbeiter"
_b.append(["Name", "Abteilung", "Eintritt", "Gehalt", "Stunden"])
_daten = [
    ("Anna Müller", "IT", 2019, 4800, 40),
    ("Bernd Schmidt", "Marketing", 2021, 3900, 32),
    ("Clara Weiß", "IT", 2017, 5600, 40),
    ("David Braun", "Vertrieb", 2022, 4200, 40),
    ("Eva Krüger", "IT", 2020, 5100, 30),
    ("Felix Wagner", "Marketing", 2018, 4400, 40),
    ("Greta Hoff", "Vertrieb", 2023, 3700, 25),
    ("Hans Berg", "IT", 2015, 6200, 40),
]
for zeile in _daten:
    _b.append(list(zeile))
_mappe.save(UEB / "personal.xlsx")


# ══════════════════════════════════════════════════════════════════
# AUFGABE 1 🟢 - Lesen
# ══════════════════════════════════════════════════════════════════
# Lade UEB/"personal.xlsx" und gib aus:
#   a) Alle Blattnamen
#   b) Anzahl Zeilen und Spalten
#   c) Den Wert von A1
#   d) Alle Namen (Spalte A, ab Zeile 2)

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 2 🟢 - Zeilen durchlaufen
# ══════════════════════════════════════════════════════════════════
# Gib alle Mitarbeiter als formatierte Tabelle im Terminal aus.
# 💡 Tipp: iter_rows(min_row=2, values_only=True)

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 3 🟡 - Auswerten
# ══════════════════════════════════════════════════════════════════
# Berechne und gib aus:
#   a) Gesamtgehaltskosten
#   b) Durchschnittsgehalt
#   c) Gehalt pro Abteilung (Summe und Durchschnitt)
#   d) Wer verdient am meisten/wenigsten?
#   e) Wie viele arbeiten Teilzeit (< 40 Stunden)?
#   f) Stundenlohn für jeden (Gehalt / (Stunden * 4.33))

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 4 🟡 - Neue Datei schreiben
# ══════════════════════════════════════════════════════════════════
# Erzeuge "auswertung.xlsx" mit einem Blatt "Abteilungen":
#   Abteilung | Anzahl | Gesamtgehalt | Ø Gehalt
# Formatiere: Kopfzeile fett + farbig, Euroformat, Spaltenbreiten.

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 5 🟡 - Spalte ergänzen
# ══════════════════════════════════════════════════════════════════
# Öffne personal.xlsx, füge zwei Spalten hinzu:
#   F: "Jahre dabei"  (2026 - Eintritt)
#   G: "Jahresgehalt" (Gehalt * 12) - als EXCEL-FORMEL!
# Formatiere G im Euroformat und speichere als "personal_erweitert.xlsx".
# ⚠️ NICHT das Original überschreiben!

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 6 🔴 - Bedingte Formatierung von Hand
# ══════════════════════════════════════════════════════════════════
# Erzeuge "gehaltsuebersicht.xlsx", in der jede Zeile eingefärbt wird:
#   Gehalt > 5000  -> grüner Hintergrund
#   Gehalt < 4000  -> roter Hintergrund
#   sonst          -> gelb
# Plus: Zebrastreifen ausschalten, Kopfzeile fixieren, Autofilter setzen.
# 💡 blatt.auto_filter.ref = blatt.dimensions

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 7 🔴 - Mehrere Dateien zusammenführen
# ══════════════════════════════════════════════════════════════════
# a) Erzeuge 4 Dateien "quartal_1.xlsx" bis "quartal_4.xlsx",
#    jede mit den Spalten Produkt|Umsatz und 5 Zufallszeilen
# b) Führe sie in "jahresuebersicht.xlsx" zusammen mit Spalte "Quartal"
# c) Ergänze ein Blatt "Zusammenfassung" mit Umsatz je Quartal
# 💡 Das ist DIE typische Büro-Aufgabe! 🎯

# 👉 Dein Code:



# ══════════════════════════════════════════════════════════════════
# AUFGABE 8 ⭐ BONUS - Diagramm
# ══════════════════════════════════════════════════════════════════
# Ergänze in "auswertung.xlsx" ein Balkendiagramm
# "Gehaltskosten pro Abteilung".
# 💡 from openpyxl.chart import BarChart, Reference

# 👉 Dein Code:



print(f"\n✅ Deine Dateien liegen in: {UEB}")
