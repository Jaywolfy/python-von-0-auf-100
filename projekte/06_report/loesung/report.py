"""
📈 Report-Generator - Musterlösung Projekt 6

Liest 12 monatliche CSV-Dateien ein und erzeugt daraus einen
formatierten Excel-Jahresbericht mit Diagrammen.

Benutzt Modul 00-24: CSV, Funktionen, Exceptions, openpyxl.

AUFRUF:
    python report.py
"""
import csv
import random
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl fehlt. Installieren mit:\n   pip install openpyxl\n")
    raise SystemExit(0)

# ====================================================================
# EINSTELLUNGEN
# ====================================================================
BASIS = Path(__file__).parent
EINGABE = BASIS / "daten"
AUSGABE = BASIS / "berichte"
EINGABE.mkdir(exist_ok=True)
AUSGABE.mkdir(exist_ok=True)

JAHR = 2026
BREITE = 54
KOPF_BLAU, KOPF_GRUEN, KOPF_ORANGE = "4472C4", "70AD47", "ED7D31"

MONATSNAMEN = ["Januar", "Februar", "März", "April", "Mai", "Juni",
               "Juli", "August", "September", "Oktober", "November", "Dezember"]

PRODUKTE = [
    ("Laptop Pro 15", "Technik", 1299.00), ("Laptop Air 13", "Technik", 899.00),
    ("Monitor 27\"", "Technik", 349.00), ("Tastatur Mech.", "Technik", 129.00),
    ("Maus Ergo", "Technik", 79.00), ("Webcam HD", "Technik", 89.00),
    ("Schreibtisch", "Möbel", 449.00), ("Bürostuhl Ergo", "Möbel", 329.00),
    ("Rollcontainer", "Möbel", 189.00), ("Regal Modular", "Möbel", 149.00),
    ("Notizbuch A4", "Bürobedarf", 12.90), ("Ordner-Set", "Bürobedarf", 24.50),
    ("Druckerpapier", "Bürobedarf", 34.90), ("Stifte-Set", "Bürobedarf", 18.90),
]


# ====================================================================
# TESTDATEN
# ====================================================================
def erzeuge_rohdaten():
    """Erzeugt 12 monatliche CSV-Dateien mit realistischen Verkaufsdaten."""
    random.seed(2026)
    saison = [0.8, 0.75, 0.9, 0.95, 1.0, 0.95, 0.85, 0.8, 1.05, 1.1, 1.35, 1.25]

    for monat in range(1, 13):
        zeilen = ["Datum;Produkt;Kategorie;Menge;Einzelpreis"]
        anzahl = int(random.randint(120, 170) * saison[monat - 1])

        for _ in range(anzahl):
            produkt, kategorie, preis = random.choice(PRODUKTE)
            tag = random.randint(1, 28)
            menge = random.randint(1, 8)
            preis_text = f"{preis:.2f}".replace(".", ",")
            zeilen.append(f"{tag:02d}.{monat:02d}.{JAHR};{produkt};"
                          f"{kategorie};{menge};{preis_text}")

        # Absichtlich fehlerhafte Zeilen
        if monat in (3, 7):
            zeilen.insert(10, f"15.{monat:02d}.{JAHR};Kaputt;Technik;abc;99,00")
        if monat == 9:
            zeilen.insert(5, "unvollstaendige;Zeile")

        pfad = EINGABE / f"umsatz_{JAHR}_{monat:02d}.csv"
        pfad.write_text("\n".join(zeilen) + "\n", encoding="utf-8")


# ====================================================================
# HILFSFUNKTIONEN
# ====================================================================
def deutsche_zahl(text):
    """Wandelt '1.234,56' in 1234.56 um."""
    return float(text.strip().replace(".", "").replace(",", "."))


def euro(betrag):
    """Formatiert einen Betrag im deutschen Format."""
    return f"{betrag:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " €"


def kopfzeile(text):
    """Gibt eine Überschrift mit Rahmen aus."""
    print("╔" + "═" * BREITE + "╗")
    print("║" + text.center(BREITE) + "║")
    print("╚" + "═" * BREITE + "╝")


def abschnitt(text):
    """Gibt eine Abschnittsüberschrift aus."""
    print(f"\n── {text} " + "─" * max(0, BREITE - len(text) - 4))


def formatiere_kopf(blatt, farbe=KOPF_BLAU):
    """Formatiert die erste Zeile als Kopfzeile."""
    for zelle in blatt[1]:
        if zelle.value is not None:
            zelle.font = Font(bold=True, color="FFFFFF", size=11)
            zelle.fill = PatternFill("solid", fgColor=farbe)
            zelle.alignment = Alignment(horizontal="center", vertical="center")
    blatt.row_dimensions[1].height = 22
    blatt.freeze_panes = "A2"


def breite_anpassen(blatt, zusatz=4, maximum=40):
    """Passt alle Spaltenbreiten an den Inhalt an."""
    for spalte in range(1, blatt.max_column + 1):
        laengste = max(len(str(blatt.cell(row=r, column=spalte).value or ""))
                       for r in range(1, blatt.max_row + 1))
        buchstabe = get_column_letter(spalte)
        blatt.column_dimensions[buchstabe].width = min(laengste + zusatz, maximum)


# ====================================================================
# EINLESEN
# ====================================================================
def lade_datei(pfad):
    """Liest eine einzelne CSV-Datei.

    Returns:
        (datensaetze, anzahl_fehler)
    """
    datensaetze = []
    fehler = 0

    with open(pfad, encoding="utf-8", newline="") as f:
        for zeile in csv.DictReader(f, delimiter=";"):
            try:
                menge = int(zeile["Menge"])
                preis = deutsche_zahl(zeile["Einzelpreis"])
                datum = datetime.strptime(zeile["Datum"], "%d.%m.%Y").date()
            except (ValueError, TypeError, KeyError, AttributeError):
                fehler += 1
                continue

            datensaetze.append({
                "datum": datum, "monat": datum.month,
                "produkt": zeile["Produkt"], "kategorie": zeile["Kategorie"],
                "menge": menge, "preis": preis,
                "umsatz": round(menge * preis, 2),
            })

    return datensaetze, fehler


def lade_alle(ordner, muster=f"umsatz_{JAHR}_*.csv"):
    """Liest alle passenden Dateien eines Ordners ein.

    Returns:
        (alle_datensaetze, anzahl_dateien, anzahl_fehler)

    Raises:
        FileNotFoundError: Wenn der Ordner nicht existiert.
    """
    ordner = Path(ordner)
    if not ordner.exists():
        raise FileNotFoundError(f"Ordner nicht gefunden: {ordner.absolute()}")

    alle = []
    dateien = 0
    fehler_gesamt = 0

    for pfad in sorted(ordner.glob(muster)):
        try:
            daten, fehler = lade_datei(pfad)
        except PermissionError:
            print(f"  ⚠️  Kein Zugriff auf {pfad.name} - übersprungen")
            continue
        except OSError as f_:
            print(f"  ⚠️  {pfad.name}: {f_}")
            continue

        alle.extend(daten)
        fehler_gesamt += fehler
        dateien += 1

    return alle, dateien, fehler_gesamt


# ====================================================================
# AUSWERTUNGEN
# ====================================================================
def auswertung_monate(daten):
    """Umsatz und Stückzahl je Monat."""
    monate = {}
    for d in daten:
        eintrag = monate.setdefault(d["monat"], {"umsatz": 0.0, "menge": 0})
        eintrag["umsatz"] += d["umsatz"]
        eintrag["menge"] += d["menge"]
    return dict(sorted(monate.items()))


def auswertung_kategorien(daten):
    """Umsatz je Kategorie."""
    summen = {}
    for d in daten:
        summen[d["kategorie"]] = summen.get(d["kategorie"], 0) + d["umsatz"]
    return dict(sorted(summen.items(), key=lambda p: -p[1]))


def top_produkte(daten, anzahl=10):
    """Die umsatzstärksten Produkte."""
    produkte = {}
    for d in daten:
        eintrag = produkte.setdefault(d["produkt"],
                                      {"umsatz": 0.0, "menge": 0,
                                       "kategorie": d["kategorie"]})
        eintrag["umsatz"] += d["umsatz"]
        eintrag["menge"] += d["menge"]
    sortiert = sorted(produkte.items(), key=lambda p: -p[1]["umsatz"])
    return sortiert[:anzahl]


def quartale(monate):
    """Fasst Monate zu Quartalen zusammen."""
    q = {"Q1": 0.0, "Q2": 0.0, "Q3": 0.0, "Q4": 0.0}
    for monat, werte in monate.items():
        q[f"Q{(monat - 1) // 3 + 1}"] += werte["umsatz"]
    return q


# ====================================================================
# EXCEL
# ====================================================================
def erzeuge_excel(daten, monate, kategorien, top, pfad):
    """Baut die formatierte Excel-Datei mit vier Blättern und Diagrammen.

    Raises:
        PermissionError: Wenn die Datei gerade geöffnet ist.
    """
    mappe = Workbook()

    # --- Blatt 1: Übersicht -----------------------------------------
    ueber = mappe.active
    ueber.title = "Übersicht"
    ueber["A1"] = f"JAHRESBERICHT {JAHR}"
    ueber["A1"].font = Font(bold=True, size=18, color=KOPF_BLAU)
    ueber.merge_cells("A1:D1")
    ueber["A2"] = f"Erstellt am {datetime.now():%d.%m.%Y %H:%M}"
    ueber["A2"].font = Font(italic=True, size=9, color="808080")

    gesamt = sum(d["umsatz"] for d in daten)
    stueck = sum(d["menge"] for d in daten)
    bester_monat = max(monate, key=lambda m: monate[m]["umsatz"])
    q = quartale(monate)
    bestes_q = max(q, key=q.get)

    kennzahlen = [
        ("Gesamtumsatz", gesamt),
        ("Ø pro Monat", gesamt / len(monate) if monate else 0),
        ("Bester Monat", monate[bester_monat]["umsatz"]),
        ("Bestes Quartal", q[bestes_q]),
        ("Artikel verkauft", stueck),
        ("Datensätze", len(daten)),
    ]
    ueber["A4"] = "Kennzahl"
    ueber["B4"] = "Wert"
    for i, (name, wert) in enumerate(kennzahlen, start=5):
        ueber.cell(row=i, column=1, value=name)
        z = ueber.cell(row=i, column=2, value=round(wert, 2))
        if "Artikel" not in name and "Datensätze" not in name:
            z.number_format = '#,##0.00 "€"'
        else:
            z.number_format = "#,##0"

    ueber["C5"] = f"Bester Monat: {MONATSNAMEN[bester_monat - 1]}"
    ueber["C6"] = f"Bestes Quartal: {bestes_q}"

    for zelle in ueber[4]:
        if zelle.value:
            zelle.font = Font(bold=True, color="FFFFFF")
            zelle.fill = PatternFill("solid", fgColor=KOPF_BLAU)
    breite_anpassen(ueber)

    # --- Blatt 2: Nach Monat ----------------------------------------
    blatt_m = mappe.create_sheet("Nach Monat")
    blatt_m.append(["Monat", "Umsatz", "Stückzahl", "Anteil"])
    for monat in range(1, 13):
        werte = monate.get(monat, {"umsatz": 0.0, "menge": 0})
        blatt_m.append([MONATSNAMEN[monat - 1], round(werte["umsatz"], 2),
                        werte["menge"],
                        werte["umsatz"] / gesamt if gesamt else 0])
    for r in range(2, blatt_m.max_row + 1):
        blatt_m.cell(row=r, column=2).number_format = '#,##0.00 "€"'
        blatt_m.cell(row=r, column=4).number_format = "0.0%"
    formatiere_kopf(blatt_m, KOPF_GRUEN)
    breite_anpassen(blatt_m)

    linie = LineChart()
    linie.title = f"Umsatzverlauf {JAHR}"
    linie.y_axis.title = "Euro"
    linie.height, linie.width = 9, 18
    werte = Reference(blatt_m, min_col=2, min_row=1, max_row=13)
    kat = Reference(blatt_m, min_col=1, min_row=2, max_row=13)
    linie.add_data(werte, titles_from_data=True)
    linie.set_categories(kat)
    blatt_m.add_chart(linie, "F2")

    # --- Blatt 3: Nach Kategorie ------------------------------------
    blatt_k = mappe.create_sheet("Nach Kategorie")
    blatt_k.append(["Kategorie", "Umsatz", "Anteil"])
    for kategorie, umsatz in kategorien.items():
        blatt_k.append([kategorie, round(umsatz, 2),
                        umsatz / gesamt if gesamt else 0])
    for r in range(2, blatt_k.max_row + 1):
        blatt_k.cell(row=r, column=2).number_format = '#,##0.00 "€"'
        blatt_k.cell(row=r, column=3).number_format = "0.0%"
    formatiere_kopf(blatt_k, KOPF_ORANGE)
    breite_anpassen(blatt_k)

    balken = BarChart()
    balken.title = "Umsatz nach Kategorie"
    balken.height, balken.width = 8, 14
    werte = Reference(blatt_k, min_col=2, min_row=1, max_row=blatt_k.max_row)
    kat = Reference(blatt_k, min_col=1, min_row=2, max_row=blatt_k.max_row)
    balken.add_data(werte, titles_from_data=True)
    balken.set_categories(kat)
    blatt_k.add_chart(balken, "E2")

    # --- Blatt 4: Top-Produkte --------------------------------------
    blatt_t = mappe.create_sheet("Top-Produkte")
    blatt_t.append(["Rang", "Produkt", "Kategorie", "Stückzahl", "Umsatz"])
    for rang, (produkt, werte) in enumerate(top, start=1):
        blatt_t.append([rang, produkt, werte["kategorie"],
                        werte["menge"], round(werte["umsatz"], 2)])
    for r in range(2, blatt_t.max_row + 1):
        blatt_t.cell(row=r, column=5).number_format = '#,##0.00 "€"'
    formatiere_kopf(blatt_t)
    breite_anpassen(blatt_t)

    # --- Blatt 5: Rohdaten ------------------------------------------
    blatt_r = mappe.create_sheet("Rohdaten")
    blatt_r.append(["Datum", "Produkt", "Kategorie", "Menge",
                    "Einzelpreis", "Umsatz"])
    for d in daten[:2000]:            # Excel-Datei nicht sprengen
        blatt_r.append([d["datum"].strftime("%d.%m.%Y"), d["produkt"],
                        d["kategorie"], d["menge"], d["preis"], d["umsatz"]])
    for r in range(2, blatt_r.max_row + 1):
        blatt_r.cell(row=r, column=5).number_format = '#,##0.00 "€"'
        blatt_r.cell(row=r, column=6).number_format = '#,##0.00 "€"'
    formatiere_kopf(blatt_r)
    breite_anpassen(blatt_r)
    blatt_r.auto_filter.ref = blatt_r.dimensions

    mappe.save(pfad)
    return pfad


def schreibe_zusammenfassung(daten, monate, kategorien, top, pfad):
    """Schreibt die Textzusammenfassung."""
    gesamt = sum(d["umsatz"] for d in daten)
    bester = max(monate, key=lambda m: monate[m]["umsatz"])

    zeilen = [
        f"JAHRESBERICHT {JAHR}",
        "=" * 56,
        f"Erstellt:         {datetime.now():%d.%m.%Y %H:%M}",
        f"Datensätze:       {len(daten):,}".replace(",", "."),
        "",
        "KENNZAHLEN",
        "-" * 56,
        f"Gesamtumsatz:     {euro(gesamt):>20}",
        f"Ø pro Monat:      {euro(gesamt / len(monate)):>20}",
        f"Bester Monat:     {MONATSNAMEN[bester - 1]:<12}"
        f"{euro(monate[bester]['umsatz']):>18}",
        f"Artikel verkauft: {sum(d['menge'] for d in daten):>20,}".replace(",", "."),
        "",
        "UMSATZ NACH KATEGORIE",
        "-" * 56,
    ]
    for kategorie, umsatz in kategorien.items():
        anteil = umsatz / gesamt if gesamt else 0
        zeilen.append(f"{kategorie:<16}{euro(umsatz):>18}  {anteil:>6.1%}")

    zeilen += ["", "TOP 5 PRODUKTE", "-" * 56]
    for rang, (produkt, werte) in enumerate(top[:5], start=1):
        zeilen.append(f"{rang}. {produkt:<22}{euro(werte['umsatz']):>18}  "
                      f"({werte['menge']} Stück)")

    pfad.write_text("\n".join(zeilen), encoding="utf-8")
    return pfad


def zeige_konsole(daten, monate, kategorien, dateien, fehler):
    """Gibt die Auswertung im Terminal aus."""
    gesamt = sum(d["umsatz"] for d in daten)
    q = quartale(monate)
    bestes_q = max(q, key=q.get)
    bester_monat = max(monate, key=lambda m: monate[m]["umsatz"])

    print(f"\n📂 Eingelesen:   {dateien} Dateien, {len(daten):,} Datensätze"
          .replace(",", "."))
    if fehler:
        print(f"⚠️  Übersprungen: {fehler} fehlerhafte Zeilen")

    abschnitt("JAHRESÜBERSICHT")
    print(f"  {'Gesamtumsatz':<26}{euro(gesamt):>26}")
    print(f"  {'Bestes Quartal':<14}{bestes_q:<12}{euro(q[bestes_q]):>26}")
    print(f"  {'Bester Monat':<14}{MONATSNAMEN[bester_monat - 1]:<12}"
          f"{euro(monate[bester_monat]['umsatz']):>26}")
    print(f"  {'Ø pro Monat':<26}{euro(gesamt / len(monate)):>26}")
    print(f"  {'Artikel verkauft':<26}{sum(d['menge'] for d in daten):>26,}"
          .replace(",", "."))

    abschnitt("UMSATZ NACH KATEGORIE")
    for kategorie, umsatz in kategorien.items():
        anteil = umsatz / gesamt if gesamt else 0
        print(f"  {kategorie:<14}{euro(umsatz):>16}  "
              f"{'█' * int(anteil * 16):<16} {anteil:>4.0%}")

    abschnitt("MONATSVERLAUF")
    hoechster = max(m["umsatz"] for m in monate.values()) or 1
    for monat in range(1, 13):
        werte = monate.get(monat, {"umsatz": 0.0})
        laenge = int(werte["umsatz"] / hoechster * 18)
        krone = "  🏆" if monat == bester_monat else ""
        print(f"  {MONATSNAMEN[monat - 1][:3]}  {werte['umsatz']:>10,.0f} €  "
              f"{'█' * laenge}{krone}".replace(",", "."))


# ====================================================================
# HAUPTPROGRAMM
# ====================================================================
def main():
    """Hauptprogramm."""
    start = time.time()
    kopfzeile(f"📈  REPORT-GENERATOR  {JAHR}")

    if not list(EINGABE.glob(f"umsatz_{JAHR}_*.csv")):
        erzeuge_rohdaten()
        print(f"\n  ℹ️  12 Rohdaten-Dateien erzeugt in {EINGABE.name}/")

    try:
        daten, dateien, fehler = lade_alle(EINGABE)
    except FileNotFoundError as f:
        print(f"\n  ❌ {f}")
        return 1

    if not daten:
        print("\n  📭 Keine gültigen Datensätze gefunden.")
        return 1

    monate = auswertung_monate(daten)
    kategorien = auswertung_kategorien(daten)
    top = top_produkte(daten)

    zeige_konsole(daten, monate, kategorien, dateien, fehler)

    excel_pfad = AUSGABE / f"jahresbericht_{JAHR}.xlsx"
    try:
        erzeuge_excel(daten, monate, kategorien, top, excel_pfad)
    except PermissionError:
        print(f"\n  ❌ {excel_pfad.name} ist geöffnet - bitte Excel schließen!")
        return 1

    text_pfad = schreibe_zusammenfassung(daten, monate, kategorien, top,
                                         AUSGABE / "zusammenfassung.txt")

    dauer = time.time() - start
    print(f"\n✅ {excel_pfad.name}  (5 Blätter, 2 Diagramme)")
    print(f"✅ {text_pfad.name}")
    print(f"⏱️  Laufzeit: {dauer:.1f} Sekunden")
    print(f"\n📁 Ausgabe: {AUSGABE}")
    print("\n💡 Von Hand hätte das ca. 3 Stunden gedauert. Jeden Monat. 🎉\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
